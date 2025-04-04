import streamlit as st
import pandas as pd
import numpy as np
from random import choices, randint
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 配置数据分布
distributions = {
    # 基本信息
    "家长身份": ["母亲", "父亲", "祖父母/外祖父母", "其他亲属"],
    "家长身份_权重": [68, 25, 5, 2],
    "年龄": ["25岁及以下", "26-30岁", "31-35岁", "36-40岁", "41岁及以上"],
    "年龄_权重": [3, 7, 40, 30, 20],
    "学历水平": ["初中及以下", "高中/中专", "大专", "本科", "硕士及以上"],
    "学历水平_权重": [5, 20, 25, 45, 5],
    "职业类型": ["公务员/事业单位", "教师", "企业职员", "自由职业", "全职家长", "个体户", "其他"],
    "职业类型_权重": [10, 15, 35, 10, 25, 5, 15],
    "家庭月收入": ["3000元及以下", "3000-5000元", "5000-8000元", "8000-12000元", "12000元及以上"],
    "家庭月收入_权重": [20, 20, 40, 20, 10],
    "家庭结构": ["核心家庭", "三代同堂", "单亲家庭", "其他"],
    "家庭结构_权重": [65, 25, 8, 2],
    "孩子年龄": ["3-4周岁", "4-5周岁", "5-6周岁", "其他"],
    "孩子年龄_权重": [40, 30, 20, 10],
    
    # 教育观念
    "能力培养": ["创造力", "规则意识", "情绪管理", "知识学习", "运动能力", "艺术兴趣", "其他"],
    "能力培养_权重": [52, 48, 45, 35, 28, 15, 4],
    "长期期待": ["身心健康", "道德品质", "独立自主", "名牌大学", "高收入工作", "兴趣爱好", "其他"],
    "长期期待_权重": [78, 65, 58, 30, 25, 20, 5],
    "游戏重要性": [1, 2, 3, 4, 5],
    "游戏重要性_权重": [2, 4, 14, 35, 45],
    "自由探索": [1, 2, 3, 4, 5],
    "自由探索_权重": [1, 3, 10, 30, 56],
    "提前学习": [1, 2, 3, 4, 5],
    "提前学习_权重": [15, 25, 30, 20, 10],
    "纠错方式": ["耐心讲道理", "共同讨论", "批评责备", "冷处理"],
    "纠错方式_权重": [65, 58, 12, 8],
    "教育方法": ["以身作则", "制定规则", "游戏引导", "物质奖励"],
    "教育方法_权重": [72, 60, 48, 18],
    "适度惩罚": ["完全必要", "有时必要", "视情况", "不太必要", "完全不必"],
    "适度惩罚_权重": [10, 65, 15, 5, 5],
    "教育内容": ["社会情感", "生活实践", "自然科学", "学科知识"],
    "教育内容_权重": [68, 55, 42, 30],
    "学科游戏": [1, 2, 3, 4, 5],
    "学科游戏_权重": [5, 10, 20, 25, 40],
    "互动时间": [1, 2, 3, 4, 5],
    "互动时间_权重": [5, 10, 20, 20, 45],
    "冲突解决": ["父母决定", "共同协商", "各执己见", "孩子决定"],
    "冲突解决_权重": [25, 55, 15, 5],
    "焦虑程度": [1, 2, 3, 4, 5],
    "焦虑程度_权重": [5, 10, 25, 30, 30],
    "焦虑来源": ["幼小衔接", "时间不足", "兴趣班", "表现不佳"],
    "焦虑来源_权重": [62, 48, 35, 12],
    "知识来源": ["书籍", "幼儿园", "自媒体", "亲友", "自我摸索"],
    "知识来源_权重": [40, 30, 55, 25, 20],
    "代际冲突": ["经常", "偶尔", "很少", "从未"],
    "代际冲突_权重": [20, 50, 25, 5],
    "传统观念": ["学而优则仕", "严父慈母", "棍棒教育"],
    "传统观念_权重": [30, 25, 10],
    "自媒体使用": ["<1小时", "1-3小时", "3-5小时", ">5小时"],
    "自媒体使用_权重": [30, 40, 20, 10],
    "家长群影响": ["非常大", "较大", "一般", "较小", "无影响"],
    "家长群影响_权重": [20, 30, 35, 10, 5]
}

def generate_data(num_records):
    data = []
    for i in range(num_records):
        record = {
            "序号": i+1,
            "提交时间": (datetime.now() - timedelta(minutes=randint(0, 1440))).strftime("%Y-%m-%d %H:%M:%S"),
            "所用时间(秒)": randint(30, 300),
            "来源": "网页",
            "来源详情": "直接访问",
            "IP地址": "潮州",
            
            # 基本信息
            "家长身份": choices(distributions["家长身份"], weights=distributions["家长身份_权重"])[0],
            "年龄": choices(distributions["年龄"], weights=distributions["年龄_权重"])[0],
            "学历水平": choices(distributions["学历水平"], weights=distributions["学历水平_权重"])[0],
            "职业类型": choices(distributions["职业类型"], weights=distributions["职业类型_权重"])[0],
            "家庭月收入": choices(distributions["家庭月收入"], weights=distributions["家庭月收入_权重"])[0],
            "家庭结构": choices(distributions["家庭结构"], weights=distributions["家庭结构_权重"])[0],
            "孩子年龄": "|".join(choices(distributions["孩子年龄"], weights=distributions["孩子年龄_权重"], k=randint(1,2))),
            
            # 教育观念
            "能力培养偏好": "|".join(choices(distributions["能力培养"], weights=distributions["能力培养_权重"], k=3)),
            "长期发展期待": "|".join(choices(distributions["长期期待"], weights=distributions["长期期待_权重"], k=3)),
            "游戏重要性评分": choices(distributions["游戏重要性"], weights=distributions["游戏重要性_权重"])[0],
            "自由探索评分": choices(distributions["自由探索"], weights=distributions["自由探索_权重"])[0],
            "提前学习评分": choices(distributions["提前学习"], weights=distributions["提前学习_权重"])[0],
            "纠错方式": "|".join(choices(distributions["纠错方式"], weights=distributions["纠错方式_权重"], k=2)),
            "常用教育方法": "|".join(choices(distributions["教育方法"], weights=distributions["教育方法_权重"], k=2)),
            "适度惩罚态度": choices(distributions["适度惩罚"], weights=distributions["适度惩罚_权重"])[0],
            "教育内容重点": "|".join(choices(distributions["教育内容"], weights=distributions["教育内容_权重"], k=2)),
            "学科游戏评分": choices(distributions["学科游戏"], weights=distributions["学科游戏_权重"])[0],
            "每日互动评分": choices(distributions["互动时间"], weights=distributions["互动时间_权重"])[0],
            "冲突解决模式": choices(distributions["冲突解决"], weights=distributions["冲突解决_权重"])[0],
            "教育焦虑评分": choices(distributions["焦虑程度"], weights=distributions["焦虑程度_权重"])[0],
            "焦虑来源": "|".join(choices(distributions["焦虑来源"], weights=distributions["焦虑来源_权重"], k=2)),
            
            # 影响因素
            "育儿知识来源": "|".join(choices(distributions["知识来源"], weights=distributions["知识来源_权重"], k=2)),
            "代际冲突频率": choices(distributions["代际冲突"], weights=distributions["代际冲突_权重"])[0],
            "传统观念认同": "|".join(choices(distributions["传统观念"], weights=distributions["传统观念_权重"], k=randint(0,2)) or "无",
            "自媒体使用时间": choices(distributions["自媒体使用"], weights=distributions["自媒体使用_权重"])[0],
            "家长群影响程度": choices(distributions["家长群影响"], weights=distributions["家长群影响_权重"])[0]
        }
        data.append(record)
    return pd.DataFrame(data)

def adjust_column_width(writer, sheet_name, df):
    workbook = writer.book
    worksheet = workbook[sheet_name]
    
    for idx, col in enumerate(df.columns):
        max_length = max((
            df[col].astype(str).map(len).max(),
            len(str(col))
        )
        worksheet.column_dimensions[get_column_letter(idx+1)].width = min(max_length + 2, 30)
        
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

# 网页界面
st.title("家长教育问卷数据生成器")
st.markdown("根据论文数据分布自动生成问卷数据")

num_records = st.number_input("要生成多少份数据?", min_value=1, value=312, step=100)

if st.button("生成数据"):
    df = generate_data(num_records)
    st.success(f"成功生成 {len(df)} 条记录!")
    st.dataframe(df.head())
    
    # 导出Excel
    output = st.radio("导出格式:", ["Excel (.xlsx)", "CSV (.csv)"])
    
    if output == "Excel (.xlsx)":
        with pd.ExcelWriter("survey_data.xlsx", engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='问卷数据')
            adjust_column_width(writer, '问卷数据', df)
        
        with open("survey_data.xlsx", "rb") as f:
            st.download_button(
                label="下载Excel文件",
                data=f,
                file_name="家长教育问卷数据.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        csv = df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
        st.download_button(
            label="下载CSV文件",
            data=csv,
            file_name="家长教育问卷数据.csv",
            mime="text/csv"
        )
